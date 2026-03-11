import json
import os
import sys
import asyncio
import time
import math
from datetime import datetime
import re
from rt_search import RealTimeSearch
from tel_send import tel_send as real_tel_send, tel_send_photo as real_tel_send_photo
from trade_logger import session_logger  # [이동] 전역으로 이동
# 기본 tel_send는 GUI에서 패치될 수 있으므로 별도 정의 (GUI 로그용)
def tel_send(msg, *args, **kwargs):
    """
    GUI 로그창에 HTML로 출력하고, 텔레그램으로도 전송합니다.
    [v5.7.2] real_tel_send를 한 번만 호출하도록 구조를 단순화했습니다.
    """
    if msg:
        # GUI 로그창(StreamRedirector)이 잡을 수 있도록 접두사와 함께 출력
        print(f"DEBUG_HTML_LOG: {msg}") 
    
    # 실제 텔레그램 전송 (HTML 태그 스트리핑 등은 real_tel_send 내부에서 처리)
    real_tel_send(msg, *args, **kwargs)

def log_and_tel(msg, *args, **kwargs):
    """
    GUI 로그와 텔레그램 모두에 전송 (중요 이벤트용)
    [v5.7.2] tel_send가 이미 두 곳 모두 처리하므로 중복 호출을 제거했습니다.
    """
    tel_send(msg, *args, **kwargs)
from check_n_sell import chk_n_sell
from acc_val import fn_kt00004
from market_hour import MarketHour
from get_seq import get_condition_list
from check_bal import fn_kt00001 as get_balance
from acc_val import fn_kt00004 as get_my_stocks
from check_n_buy import ACCOUNT_CACHE, load_json_safe, update_stock_condition
from acc_realized import fn_kt00006 as get_realized_pnl
from acc_diary import fn_ka10170 as get_trade_diary, fn_ka10077 as get_realized_detail, fn_ka10076 as get_exec_list
from login import fn_au10001
import pandas as pd

def _generate_pnl_chart_png(pnl_history: list, output_dir: str) -> str:
    """
    [v5.7.13] Qt-Free 수익 차트 PNG 생성기.
    matplotlib의 Agg 백엔드(화면 표시 없이 파일 저장 전용)를 사용해
    교착상태(Deadlock) 없이 안전하게 차트를 렌더링합니다.
    실패 시 빈 문자열 반환.
    """
    try:
        import matplotlib
        matplotlib.use('Agg') # Qt 전혀 건드리지 않는 안전한 백엔드
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        from datetime import datetime as dt
        
        # 데이터 유효성 검사
        if not pnl_history or len(pnl_history) < 2:
            return ""
        
        # 데이터 추출 및 정제
        times, pnls = [], []
        for item in pnl_history:
            if isinstance(item, dict) and 'time' in item and 'pnl' in item:
                try:
                    t = dt.strptime(item['time'], "%H:%M:%S")
                    times.append(t)
                    pnls.append(int(item['pnl']))
                except: pass
        
        if len(times) < 2:
            return ""
        
        # 차트 스타일 설정 (KipoStock 다크 테마)
        fig, ax = plt.subplots(figsize=(10, 3.5), facecolor='#1a1a2e')
        ax.set_facecolor('#16213e')
        
        # 수익/손실 구간 색상 분리
        final_pnl = pnls[-1]
        line_color = '#ff4444' if final_pnl >= 0 else '#33b5e5'
        
        ax.plot(times, pnls, color=line_color, linewidth=2.0, zorder=3)
        ax.fill_between(times, pnls, 0,
                        where=[p >= 0 for p in pnls], color='#ff444433', zorder=2)
        ax.fill_between(times, pnls, 0,
                        where=[p < 0 for p in pnls], color='#33b5e533', zorder=2)
        
        # 기준선 (0원)
        ax.axhline(y=0, color='#555555', linewidth=1.0, linestyle='--', zorder=1)
        
        # 최종값 레이블
        pnl_txt = f"{final_pnl:+,}원"
        txt_color = '#ff6666' if final_pnl >= 0 else '#66aaff'
        ax.annotate(pnl_txt, xy=(times[-1], pnls[-1]),
                    xytext=(5, 5), textcoords='offset points',
                    fontsize=11, color=txt_color, fontweight='bold',
                    fontproperties=_get_korean_font())
        
        # X축 시간 포맷
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
        ax.xaxis.set_major_locator(mdates.MinuteLocator(byminute=[0, 30]))
        plt.xticks(color='#aaaaaa', fontsize=9)
        plt.yticks(color='#aaaaaa', fontsize=9)
        
        # Y축 포맷 (원 단위)
        ax.yaxis.set_major_formatter(
            matplotlib.ticker.FuncFormatter(lambda x, p: f'{int(x):+,}')
        )
        
        ax.set_title(f'[실시간 누적 수익 현황]  ({times[0].strftime("%H:%M")} ~ {times[-1].strftime("%H:%M")})',
                     color='#f1c40f', fontsize=12, pad=8,
                     fontproperties=_get_korean_font())
        
        for spine in ax.spines.values():
            spine.set_edgecolor('#333355')
        
        ax.grid(True, alpha=0.15, color='#aaaaaa')
        plt.tight_layout(pad=1.0)
        
        # 파일 저장
        if not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        save_path = os.path.join(output_dir, f'pnl_chart_{ts}.png')
        plt.savefig(save_path, dpi=110, bbox_inches='tight',
                    facecolor='#1a1a2e', edgecolor='none')
        plt.close(fig)
        plt.close('all')
        return save_path
        
    except Exception as e:
        print(f"⚠️ [차트생성 실패] {e}")
        try:
            plt.close('all')
        except: pass
        return ""

def _get_korean_font():
    """한글 폰트를 안전하게 로드합니다."""
    try:
        import matplotlib.font_manager as fm
        for font_name in ['Malgun Gothic', 'NanumGothic', 'AppleGothic']:
            try:
                return fm.FontProperties(family=font_name)
            except: pass
    except: pass
    return None

class ChatCommand:
    def __init__(self):
        self.rt_search = RealTimeSearch(on_connection_closed=self._on_connection_closed)
        self.on_graph_update = None # [신규 v5.6.4] 차트 실시간 갱신 트리거 콜백
        
        # [수정] 경로 설정 로직 변경
        if getattr(sys, 'frozen', False):
            # EXE 실행 시
            self.script_dir = os.path.dirname(sys.executable)
        else:
            # 파이썬 실행 시
            self.script_dir = os.path.dirname(os.path.abspath(__file__))
            
        self.settings_file = os.path.join(self.script_dir, 'settings.json')
        self.stock_conditions_file = os.path.join(self.script_dir, 'stock_conditions.json')
        self.config_file = os.path.join(self.script_dir, 'config.py')
        self.daily_buy_times_file = os.path.join(self.script_dir, 'daily_buy_times.json')
        self.data_dir = os.path.join(self.script_dir, 'LogData')
        if not os.path.exists(self.data_dir):
            try: os.makedirs(self.data_dir)
            except: pass
        
        self.check_n_sell_task = None
        self.account_sync_task = None
        self.token = None
        self.is_starting = False # [신규] 중복 시작(R10001) 방지용 플래그
        
        # [신규] 원격/명령어 인터페이스를 위한 콜백
        self.on_clear_logs = None # [신규] GUI 로그 초기화 콜백
        self.on_request_log_file = None # [신규] 로그 파일 저장 요청 콜백
        self.on_auto_sequence = None # [신규] 시퀀스 자동 시작 콜백
        self.on_condition_loaded = None # [신규] 목록 로드 완료 콜백
        self.on_start = None # [신규] 엔진 시작 성공 콜백
        self.on_stop = None # [신규] 엔진 정지 콜백
        self.on_open_config = None # [신규 v6.1.21] 불타기 설정창 오픈 콜백
        self.on_open_ai_settings = None # [신규 v6.1.21] AI 음성 설정창 오픈 콜백
        
        # [신규] 재연결 관련 제어 변수 (v3.0 지수 백오프용)
        self.reconnect_attempts = 0
        self.max_reconnect_delay = 60 # 최대 대기 시간 (60초)
        
        # [신규] 시작/중지 요청 콜백 (GUI를 거쳐 실행되도록)
        self.on_start_request = None
        self.on_stop_request = None
        self.on_news_result = None  # [신규] 뉴스 분석 결과 GUI 전달용 콜백
        self.on_ai_voice_response = None # [신규 v5.4.0] 제미나이 AI 비서의 음성 답변 데이터 전달용 콜백
        
        # [신규] rt_search의 콜백을 wrapper로 연결
        self.rt_search.on_condition_loaded = self._on_condition_loaded_wrapper
        # [신규 v5.1] 실시간 뉴스 결과 콜백 연결
        self.rt_search.on_news_result = lambda msg: self.on_news_result(msg) if self.on_news_result else None
        # [신규] 가속도 추가 매수 콜백 등록
        self.rt_search.on_acceleration_trigger = self.on_accel_buy_trigger

        # [신규] 이전 세션 데이터 복원
        try:
            session_logger.load_session()
        except Exception as e:
            print(f"⚠️ [ChatCommand] 세션 복원 중 예외: {e}")

    def _on_condition_loaded_wrapper(self):
        if self.on_condition_loaded:
            self.on_condition_loaded()

    def get_token(self):
        """새로운 토큰을 발급받고 모든 모듈에 강제 동기화합니다."""
        try:
            token = fn_au10001()
            if token:
                self.token = token
                if self.rt_search:
                    self.rt_search.token = token
                print(f"✅ 새로운 토큰 발급 및 동기화 완료: {token[:10]}...")
                return token
            return None
        except Exception as e:
            print(f"❌ 토큰 발급 중 오류: {e}")
            return None

    async def _account_sync_loop(self):
        """계좌 정보를 메모리에 동기화하며 인증 에러 시 즉시 재시도합니다."""
        print("🔄 계좌 동기화 루프 가동 시작 (로그 최소화 모드)")
        while self.rt_search.keep_running:
            try:
                if not self.token:
                    self.get_token()
                    await asyncio.sleep(2)

                loop = asyncio.get_event_loop()
                try:
                    from check_n_buy import update_account_cache
                    await loop.run_in_executor(None, update_account_cache, self.token)
                except Exception as api_err:
                    err_msg = str(api_err)
                    if any(x in err_msg for x in ['8005', 'Token', 'entr', 'Invalid']):
                        print(f"⚠️ 인증 실패 감지: 토큰을 재발급합니다.")
                        self.get_token()
                        await asyncio.sleep(2)
                        continue 
            except Exception as e:
                print(f"⚠️ 계좌 동기화 루프 예외: {e}")
            await asyncio.sleep(1.0)

    async def _check_n_sell_loop(self):
        """매도 체크 루프"""
        failure_count = 0
        while self.rt_search.keep_running:
            try:
                if MarketHour.is_waiting_period():
                    await asyncio.sleep(0.5)
                    continue

                if not self.token: 
                    await asyncio.sleep(1)
                    continue
                    
                success = await asyncio.get_event_loop().run_in_executor(None, chk_n_sell, self.token)
                failure_count = 0 if success else failure_count + 1
                
                # [v5.5] 실시간 차트 API 동기화 감지 (매도 발생 시 1회)
                from trade_logger import session_logger
                if getattr(session_logger, 'sync_required', False):
                    session_logger.sync_required = False
                    # UI 멈춤 방지를 위해 백그라운드 태스크로 분리하여 API 실시간 동기화 호출
                    asyncio.create_task(self.today(sync_only=True))
                
                if failure_count >= 20:
                    print("⚠️ 매도 루프 연속 실패로 재시작 시도")
                    break 
                
                # [최적화] CPU 점유율 과다 방지를 위해 0.5초 대기 (초고속 성능 유지와 부하 균형)
                await asyncio.sleep(0.5) 
            except asyncio.CancelledError:
                break
            except Exception as e:
                print(f"⚠️ 매도 루프 에러: {e}")
                await asyncio.sleep(1) # 에러 시 잠시 대기
                failure_count += 1
            await asyncio.sleep(0.1)

    async def start(self, profile_info=None, manual=False):
        """시스템 시작"""
        if self.is_starting:
            print("⏳ [알람] 이미 엔진을 시작하는 중입니다. 중복 요청을 무시합니다.")
            return False
            
        try:
            self.is_starting = True
            await self._cancel_tasks()
            
            # [Fix] 중복 로그인(R10001) 방지: 기존 소켓이 열려있다면 닫고 시작
            if self.rt_search.connected or self.rt_search.websocket:
                print("🔄 [재접속] 기존 연결을 정리하고 새로 시작합니다...")
                await self.rt_search.stop()
                await asyncio.sleep(2.0) # 세션 정리 대기 시간 추가 증가 (1.5 -> 2.0)

            token = self.get_token()
            if not token:
                log_and_tel("❌ 토큰 발급 실패")
                self.is_starting = False # Ensure flag is reset on failure
                return False
            
            self.update_setting('auto_start', True)
            
            # [수정] 수동 시작(manual=True)인 경우 사용자 설정을 무시하고 실제 장 시간(09:00~15:30)만 체크
            if manual:
                if not MarketHour.is_actual_market_open_time():
                    log_and_tel(f"⚠️ [거부] 실제 장 데이터 수신 시간이 아닙니다. (수동 시작은 08:30~15:29 사이에만 가능)")
                    self.is_starting = False
                    return False
                # [신규] 수동 모드 플래그 활성화 -> is_waiting_period() 무시
                MarketHour.set_manual_mode(True)
            else:
                # [Fix] 오토 시퀀스(자동) 시작인 경우, 혹시 남아있을 수 있는 수동 모드 플래그를 확실히 해제
                MarketHour.set_manual_mode(False)
                # 일반 시퀀스 시작 등은 기존처럼 사용자 설정 시간(Waiting Period) 체크
                if MarketHour.is_waiting_period():
                    now_str = datetime.now().strftime('%H:%M:%S')
                    print(f"⚠️ [거부] 설정된 매매 시간이 아닙니다. (현재: {now_str})")
                    self.is_starting = False 
                    return False
            
            loop = asyncio.get_event_loop()
            try:
                from check_n_buy import update_account_cache
                await loop.run_in_executor(None, update_account_cache, token)
                balance_raw = ACCOUNT_CACHE.get('balance', 0)
            except Exception as e:
                print(f"⚠️ 계좌 정보 초기화 중 오류: {e} - 계속 진행합니다.")
                balance_raw = 0
            
            acnt_no = ACCOUNT_CACHE.get('acnt_no')
            success = await self.rt_search.start(token, acnt_no=acnt_no)
            if success:
                self.check_n_sell_task = asyncio.create_task(self._check_n_sell_loop())
                self.account_sync_task = asyncio.create_task(self._account_sync_loop())
                log_and_tel(f"🚀 실시간 감시 엔진 {profile_info if profile_info else '기본'} 모드 시작 완료")
                if self.on_start: self.on_start() # [신규] GUI 상태 동기화
                return True
            else:
                self.is_starting = False # Ensure flag is reset on failure
                return False
        except Exception as e:
            log_and_tel(f"❌ start 오류: {e}")
            return False
        finally:
            self.is_starting = False

    async def stop(self, set_auto_start_false=True, quiet=False):
        """시스템 중지"""
        try:
            # [Fix] 엔진 정지 시 수동 모드 플래그 무조건 해제 (다음 자동 시작을 위해)
            MarketHour.set_manual_mode(False)
            
            if set_auto_start_false:
                self.update_setting('auto_start', False)
            await self._cancel_tasks()
            await self.rt_search.stop()
            if not quiet:
                log_and_tel("⏹ 실시간 감시 엔진이 정지되었습니다.")
                if self.on_stop: self.on_stop() # [신규] GUI 상태 동기화
            return True
        except Exception as e:
            if not quiet: log_and_tel(f"❌ stop 오류: {e}")
            return False

    async def sync_marked_indices(self, marked_list):
        """[신규] GUI의 마킹 상태를 검색 엔진과 동기화"""
        if self.rt_search:
            self.rt_search.marked_indices = set(marked_list)
            # print(f"🔄 [Sync] 마킹 동기화 완료: {marked_list}")

    async def on_accel_buy_trigger(self, code, name, last_1s, avg_5m):
        """[신규] 가속도 조건 만족 시 추가 매수 실행"""
        from check_n_buy import add_buy
        # 기본 1주 추가 매수 (설정에 따라 변경 가능하도록 확장 가능)
        loop = asyncio.get_event_loop()
        loop.run_in_executor(None, add_buy, code, self.token, name, 1)

    async def _cancel_tasks(self):
        """실행 중인 태스크 취소 및 대기"""
        tasks = [('매도', self.check_n_sell_task), ('계좌', self.account_sync_task)]
        for name, task in tasks:
            if task and not task.done():
                task.cancel()
                try:
                    await task
                except asyncio.CancelledError:
                    pass
                except Exception as e:
                    print(f"⚠️ {name} 태스크 종료 중 에러: {e}")
        
        self.check_n_sell_task = None
        self.account_sync_task = None

    async def _on_connection_closed(self):
        """재연결 콜백 (v3.0 지수 백오프 적용)"""
        if self.is_starting:
            print("🔄 [안내] 엔진 재시작 중으로 자동 재연결을 건너뜁니다.")
            return

        await self.stop(set_auto_start_false=False)
        
        # 지수 백오프 대기 시간 계산: 2^attempts + random jitter (jitter는 일단 생략)
        self.reconnect_attempts += 1
        delay = min(self.max_reconnect_delay, 2 ** self.reconnect_attempts)
        
        print(f"⚠️ [재연결] 소켓 끊김 감시... {delay}초 후 재시도합니다. (시도 횟수: {self.reconnect_attempts})")
        await asyncio.sleep(delay)
        

        success = await self.start()
        if success:
            self.reconnect_attempts = 0 # 성공 시 횟수 초기화

    async def report(self, seq=None):
        """종합 누적 리포트: 당일 전체 매매 일지 + 계좌 현황 + 퀀트 지표 + 파일 저장"""
        try:
            print(f"📊 [REPORT] {'시퀀스 '+str(seq)+' 종료 후 ' if seq else ''}누적 리포트 생성 중...")
            log_and_tel("⏳ <b>리포트 데이터를 수집 중입니다. 잠시만 기다려 주세요...</b>", parse_mode='HTML', msg_type='report')
            
            diary_text, stats = await self.today(summary_only=False, return_text=True, return_stats=True)
            
            peak_time_str = stats.get('peak_pnl_time', '약속된 시간 없음') if stats else '약속된 시간 없음'
            peak_pnl_val = stats.get('peak_pnl', 0) if stats else 0
            
            if not self.token: self.get_token()
            loop = asyncio.get_event_loop()
            
            balance_res = await loop.run_in_executor(None, get_balance, 'N', '', self.token, True)
            if balance_res and isinstance(balance_res, dict):
                 balance_raw = balance_res.get('balance', 0)
            else:
                 balance_raw = balance_res
            balance_str = f"{int(balance_raw):,}원" if balance_raw else "조회 실패"
            
            account_data_raw = await loop.run_in_executor(None, fn_kt00004, False, 'N', '', self.token)
            if isinstance(account_data_raw, dict):
                account_data = account_data_raw.get('stocks', [])
            else:
                account_data = account_data_raw
            

            # 3. 종합 요약 메시지 구성 (GUI 및 텔레그램용)
            title_prefix = f"시퀀스 {seq} 종료 후 " if seq else ""
            total_pnl = stats.get('total_pnl', 0) if stats else 0
            avg_pnl_rt = stats.get('total_rt', stats.get('avg_pnl_rt', 0)) if stats else 0
            pnl_color = "#ff4444" if total_pnl >= 0 else "#33b5e5"
            peak_time_str = stats.get('peak_pnl_time', '약속된 시간 없음') if stats else '약속된 시간 없음'
            peak_pnl_val = stats.get('peak_pnl', 0) if stats else 0

            # [v5.7.13] Qt-Free Agg 백엔드로 수익 차트 PNG 생성 (교착상태 완전 해결)
            try:
                from trade_logger import session_logger
                png_path = await loop.run_in_executor(
                    None, _generate_pnl_chart_png,
                    list(session_logger.pnl_history), self.data_dir
                )
                img_html = f'<img src="file:///{png_path.replace(chr(92), "/")}" width="600"><br>' if png_path else ""
                if png_path:
                    print(f"✅ [차트] 수익 차트 생성 완료: {os.path.basename(png_path)}")
                else:
                    img_html = ""
                    print("ℹ️ [차트] 거래 데이터 부족으로 차트 생략")
            except Exception as chart_err:
                img_html = ""
                png_path = ""
                print(f"⚠️ [차트 생성 에러] {chart_err}")

            # GUI용 메시지 (구조 경량화)
            msg_gui = ""
            msg_gui += f"🚀 <b>[{title_prefix if title_prefix else '오늘 전체 '}매매 종합 리포트]</b>\n"
            msg_gui += "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            msg_gui += f"📅 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            msg_gui += f"💰 <b>당일 총손익 :</b> <font color='{pnl_color}'><b>{total_pnl:+,}원 ({avg_pnl_rt:+.2f}%)</b></font>\n"
            msg_gui += f"👑 <b>최고 수익 시간 :</b> <font color='#f1c40f'><b>{peak_time_str} ({peak_pnl_val:+,}원)</b></font>\n"
            msg_gui += "────────────────────────────────────────\n"
            
            if stats:
                wr = stats.get('win_rate', 0)
                mdd = stats.get('mdd', 0)
                pf = stats.get('profit_factor', 0)
                pr = stats.get('payoff_ratio', 0)
                ex = stats.get('expectancy', 0)
                sr = stats.get('sharpe_ratio', 0)
                
                wr_color = "#ff4444" if wr >= 70 else ("#33b5e5" if wr <= 40 else "#ffffff")
                pf_color = "#ff4444" if pf >= 2.0 else ("#33b5e5" if pf < 1.0 else "#ffffff")
                
                msg_gui += f"   📊 <b>승  률 :</b> <font color='{wr_color}'><b>{wr:.1f}%</b></font>\n"
                msg_gui += f"   💰 <b>PF(Profit Factor) :</b> <font color='{pf_color}'><b>{pf:.2f}</b></font>\n"
                msg_gui += f"   ⚖️ <b>손익비(Payoff Ratio) :</b> <b>{pr:.2f}</b>\n"
                msg_gui += f"   🎯 <b>매매 기댓값 :</b> <font color='#ffbb33'><b>{int(ex):,}원</b></font>\n"
                msg_gui += f"   📉 <b>MDD(최대낙폭) :</b> <font color='#ffbb33'><b>{int(mdd):,}원</b></font>\n"
                msg_gui += f"   📈 <b>샤프 지수 :</b> <b>{sr:.2f}</b>\n"
                msg_gui += "────────────────────────────────────────\n"

            # [v5.7.3] 렌더링 부하 최소화를 위해 전략별/조건별 통계 출력 개수 대폭 제한 (최대 5개)
            strat_stats = stats.get('strat_stats', {}) if stats else {}
            if strat_stats:
                msg_gui += "📂 <b>[ 매수 전략별 매매현황 ]</b>\n"
                sorted_strats = sorted(strat_stats.items(), key=lambda x: x[1]['pnl'], reverse=True)[:5]
                for s_key, s_data in sorted_strats:
                    s_pnl = s_data['pnl']
                    s_rt = (s_pnl / s_data['buy_amt'] * 100) if s_data['buy_amt'] > 0 else 0
                    s_color = "#ff4444" if s_pnl >= 0 else "#33b5e5"
                    msg_gui += f"   🔹 {s_data['nm']:<10}: <font color='{s_color}'><b>{s_pnl:+,}원 ({s_rt:+.2f}%)</b></font>\n"
                if len(strat_stats) > 5: msg_gui += f"   ... (기타 {len(strat_stats)-5}개 전략 생략)\n"
                msg_gui += "────────────────────────────────────────\n"

            cond_stats = stats.get('cond_stats', {}) if stats else {}
            if cond_stats or img_html:
                # [v5.7.13] 실시간 수익 차트를 조건식별 누적 통계 바로 위에 출력
                if img_html:
                    msg_gui += f"{img_html}"
                    msg_gui += "────────────────────────────────────────\n"
            if cond_stats:
                msg_gui += "📂 <b>[ 조건식별 매매현황 ]</b>\n"
                sorted_conds = sorted(cond_stats.items(), key=lambda x: x[1]['pnl'], reverse=True)[:5]
                for c_name, c_data in sorted_conds:
                    c_pnl = c_data['pnl']
                    c_rt = (c_pnl / c_data['buy_amt'] * 100) if c_data['buy_amt'] > 0 else 0
                    c_color = "#ff4444" if c_pnl >= 0 else "#33b5e5"
                    msg_gui += f"   🔹 {c_name[:8]:<8}: <font color='{c_color}'><b>{c_pnl:+,}원 ({c_rt:+.2f}%)</b></font>\n"
                if len(cond_stats) > 5: msg_gui += f"   ... (기타 {len(cond_stats)-5}개 조건 생략)\n"
                msg_gui += "────────────────────────────────────────\n"

            msg_gui += "📈 <b>[현재 보유 종목]</b>\n"
            if account_data:
                # [v5.7.3] 보유 종목도 최대 10개로 제한하여 스크롤 렌더링 부하 감소
                for s in account_data[:10]:
                    pl_rt = float(s['pl_rt'])
                    color = "#ff4444" if pl_rt > 0 else "#33b5e5"
                    msg_gui += f"   🔹 {s['stk_nm'][:6]}: <font color='{color}'>{pl_rt:+.2f}%</font>\n"
                if len(account_data) > 10: msg_gui += f"   ... (외 {len(account_data)-10}종목)\n"
            else:
                msg_gui += "   보유 중인 종목 없음\n"
            msg_gui += "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"

            # [v5.7.15] GUI 로그창 출력을 위한 핵심 코드 복구
            # [v5.7.12 크래시 수정] 수천 자 HTML을 print()로 GUI 로그창에 한번에 밀어 넣으면
            # PyQt 렌더러 스레드 블로킹 → 프로그램 다운 발생 가능성이 있으나,
            # 현재 v5.7.14의 StreamRedirector가 4000자 단위로 알아서 나눠서 쏴주므로 안전하게 출력합니다.
            print(f"DEBUG_HTML_LOG: {msg_gui}")
            print(f"✅ [REPORT] GUI 메시지 구성 완료 (길이: {len(msg_gui):,}자)")
            
            # 텔레그램 메시지 정규화
            msg_tel = re.sub(r'<div.*?</div>', '', msg_gui) # 이미지 영역 제거
            msg_tel = re.sub(r'<font.*?>', '', msg_tel).replace('</font>', '')
            msg_tel = msg_tel.replace('<br>', '\n').replace('<br/>', '\n')
            
            if png_path and os.path.exists(png_path):
                try:
                    real_tel_send_photo(png_path, caption=f"📊 {title_prefix if title_prefix else '오늘 전체 '}수익 차트")
                except: pass
            
            real_tel_send(msg_tel, parse_mode='HTML', msg_type='report')
            
            # 4. 전체 리포트 텍스트 파일 저장 (TXT)
            try:
                date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                # HTML 태그 제거 및 평문 텍스트 리포트 생성
                txt_report = f"==== [ KipoStock 통합 리포트 ] {date_str} ====\n\n"
                txt_report += "1. 당일 매매 내역 (Trade Diary)\n"
                txt_report += "--------------------------------------------------\n"
                # [v5.7.3 프리징 파훼법 핵심2] 무거운 정규식(Catastrophic Backtracking) 제거 
                # 수천 줄 데이터를 다룰 때 <[^>]*> 패턴은 메인 스레드 락을 유발합니다. 단순 replace로 경량화.
                clean_diary = diary_text.replace('<b>', '').replace('</b>', '').replace('\n<br>', '\n').replace('<br>', '\n') if diary_text else "(거래 내역 없음)\n"
                clean_diary = re.sub(r'<font.*?>', '', clean_diary).replace('</font>', '')
                txt_report += clean_diary
                
                txt_report += "\n\n2. 계좌 현황 및 세션 요약\n"
                txt_report += "--------------------------------------------------\n"
                # [v5.7.3 Bugfix] msg_gui 기반으로 텍스트 추출 (기존의 정의되지 않은 msg 변수 오류 방지)
                clean_summary = msg_gui.replace('<b>', '').replace('</b>', '').replace('<br>', '\n').replace('<br/>', '\n')
                clean_summary = re.sub(r'<font.*?>', '', clean_summary).replace('</font>', '')
                clean_summary = re.sub(r'<div.*?</div>', '', clean_summary) # 차트 이미지 html 블럭 제거
                txt_report += clean_summary
                
                # 로그 디렉토리 확인 및 저장
                log_dir = self.data_dir
                if not os.path.exists(log_dir): os.makedirs(log_dir)
                
                ts_file = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"FullReport_{ts_file}.txt"
                save_path = os.path.join(log_dir, filename)
                
                with open(save_path, 'w', encoding='utf-8') as f:
                    f.write(txt_report)
                
                log_and_tel(f"<font color='#28a745'>💾 <b>종합 리포트 파일 저장 완료:</b> {filename}</font>", parse_mode='HTML', msg_type='report')
                print(f"✅ 리포트 파일 저장 완료: {save_path}")
            except Exception as fe:
                print(f"❌ 리포트 파일 저장 중 오류: {fe}")

            return True
        except Exception as e:
            print(f"❌ [REPORT] 오류: {e}")
            log_and_tel(f"❌ <b>리포트 생성 중 오류 발생:</b> {e}", parse_mode='HTML')
            return False

    def get_pnl_timeline(self):
        """실시간 수익 그래프를 위한 누적 손익 타임라인 데이터를 반환합니다."""
        try:
            from trade_logger import session_logger
            # session_logger.pnl_history는 [{'time': 'HH:MM:SS', 'pnl': 123}, ...] 형식
            if not hasattr(session_logger, 'pnl_history') or not session_logger.pnl_history:
                return []
            
            # 리포트와 동일하게 누적 합산 로직을 보장하되, 실시간 앱 데이터를 우선함
            return session_logger.pnl_history
        except Exception as e:
            print(f"⚠️ P&L 타임라인 추출 실패: {e}")
            return []

    async def today(self, sort_mode=None, is_reverse=False, summary_only=False, send_telegram=False, return_text=False, return_stats=False, sync_only=False):
        """당일 매매 일지 조회 (Hybrid: ka10170 전체목록 + ka10077 상세세금 + ka10076 체결시간복원 + V5.5 API동기화)"""
        if not sync_only:
            print(f"▶ Today 명령어 수신 (모드: {sort_mode}, 역순: {is_reverse}, 요약: {summary_only}, 텔레그램전송: {send_telegram})")
        try:
            if not self.token:
                self.get_token()
                
            loop = asyncio.get_event_loop()
            
            res_list = await loop.run_in_executor(None, get_trade_diary, self.token)
            diary_list = res_list.get('list', [])
            
            if not diary_list:
                tel_send("📭 오늘 매매 내역이 없습니다.")
                return

            exec_time_map = {}
            try:
                exec_data = await loop.run_in_executor(None, get_exec_list, self.token)
                if exec_data:
                    for ex in exec_data:
                        t_code = ex.get('stk_cd', '').replace('A', '')
                        t_time = ex.get('ord_tmd', '') 
                        t_type = ex.get('sell_buy_tp_code', '') 
                        
                        if t_code and t_time and t_type == '2':
                            current_min = exec_time_map.get(t_code, "999999")
                            if t_time < current_min:
                                exec_time_map[t_code] = t_time
            except Exception as ex_err:
                print(f"⚠️ [TimeRestore] 체결내역 조회 실패: {ex_err}")

            cond_mapping = {}
            mapping_file = self.stock_conditions_file
            cond_mapping = load_json_safe(mapping_file)

            bt_data = {}
            try:
                bt_path = self.daily_buy_times_file
                bt_data = load_json_safe(bt_path)
            except: pass

            processed_data = []
            for item in diary_list:
                try:
                    code = item['stk_cd'].replace('A', '')
                    def val(keys):
                        for k in keys:
                            v = item.get(k)
                            if v is not None and str(v).strip() != "": return v
                        return 0

                    mapping_val = cond_mapping.get(code, "직접매매")
                    cond_name = "직접매매"
                    strat_key = "none"
                    strat_nm = "--"
                    
                    found_buy_time_entry = bt_data.get(code)
                    found_buy_time = found_buy_time_entry.get('time') if isinstance(found_buy_time_entry, dict) else found_buy_time_entry
                    
                    if isinstance(mapping_val, dict):
                        cond_name = mapping_val.get('name', "직접매매")
                        strat_key = mapping_val.get('strat', 'none')
                        strat_map = {'qty': '1주', 'amount': '금액', 'percent': '비율', 'HTS': 'HTS'}
                        strat_nm = strat_map.get(strat_key, '--')
                        
                        if not found_buy_time:
                            found_buy_time = mapping_val.get('time')
                    else:
                        cond_name = str(mapping_val)
                    
                    is_restored = False
                    if not found_buy_time:
                         api_time = exec_time_map.get(code)
                         if api_time and len(api_time) == 6:
                             found_buy_time = f"{api_time[:2]}:{api_time[2:4]}:{api_time[4:]}"
                             is_restored = True
                             if strat_nm == '--': 
                                 strat_nm = "HTS"
                                 strat_key = "HTS" # [신규] 통계용 키 명시
                                 cond_name = "외부체결(복원)"
                                 
                                 # [v6.0.8] 외부(HTS) 체결 시간 복원 시 장부에 영구 저장 (불타기 연동 핵심)
                                 update_stock_condition(
                                     code, 
                                     name=cond_name, 
                                     strat=strat_key, 
                                     time_val=found_buy_time
                                 )

                    current_time_str = datetime.now().strftime("%H:%M:%S")
                    is_overnight = False
                    
                    buy_amt_val = int(float(val(['buy_amt', 'tot_buy_amt'])))
                    
                    if not is_restored:
                        if buy_amt_val <= 0 or (found_buy_time and found_buy_time > current_time_str):
                            is_overnight = True
                    
                    final_buy_time = found_buy_time if found_buy_time else "99:99:99"
                    if is_overnight:
                        if found_buy_time: final_buy_time = f"전일 {found_buy_time[:5]}"
                        else: final_buy_time = "[전일]"
                    
                    row = {
                        'code': code,
                        'name': item.get('stk_nm', '--'),
                        'buy_time': final_buy_time,
                        'buy_avg': int(float(val(['buy_avg_pric', 'buy_avg_prc']))),
                        'buy_qty': int(float(val(['buy_qty', 'tot_buy_qty']))),
                        'buy_amt': buy_amt_val,
                        'sel_avg': int(float(val(['sel_avg_pric', 'sel_avg_prc', 'sell_avg_pric']))),
                        'sel_qty': int(float(val(['sell_qty', 'tot_sel_qty', 'sel_qty']))),
                        'sel_amt': int(float(val(['sell_amt', 'tot_sel_amt', 'sel_amt']))),
                        'tax': int(float(val(['cmsn_alm_tax', 'cmsn_tax', 'tax']))),
                        'pnl': int(float(val(['pl_amt', 'pnl_amt', 'rznd_pnl', 'tdy_sel_pl']))),
                        'pnl_rt': float(val(['prft_rt', 'pl_rt', 'profit_rate'])),
                        'cond_name': cond_name,
                        'strat_key': strat_key,
                        'strat_nm': strat_nm,
                        'is_overnight': is_overnight
                    }
                    processed_data.append(row)
                except Exception as e: 
                    print(f"⚠️ [Today] 종목({item.get('stk_nm')}) 파싱 스킵: {e}")
                    continue

            if sort_mode == 'jun':
                processed_data.sort(key=lambda x: x['strat_nm'], reverse=is_reverse)
            elif sort_mode == 'sic':
                processed_data.sort(key=lambda x: x['cond_name'], reverse=is_reverse)
            elif sort_mode == 'son':
                processed_data.sort(key=lambda x: x['pnl'], reverse=not is_reverse) 
            else:
                processed_data.sort(key=lambda x: x['buy_time'], reverse=is_reverse)

            total_b_amt = sum(r['buy_amt'] for r in processed_data)
            total_s_amt = sum(r['sel_amt'] for r in processed_data)
            total_tax = sum(r['tax'] for r in processed_data)
            total_pnl = sum(r['pnl'] for r in processed_data)
            
            # [v6.6.0] API의 전체 손익(total section)과 리스트 합산 비교 및 보정
            api_total_pnl = int(float(res_list.get('total', {}).get('tdy_sel_pl', total_pnl)))
            if abs(api_total_pnl - total_pnl) > 10:
                print(f"⚖️ [Report Sync] 리포트 합산({total_pnl:,})과 API 전체 손익({api_total_pnl:,}) 차이 발견 -> API 실판매 데이터로 보정합니다.")
                # 차액만큼을 '기타 손익' 항목으로 가상 추가하거나 total_pnl을 직접 수정
                total_pnl = api_total_pnl 
            
            count = len(processed_data)
            
            # [신규] 당일 전략별/조건식별 매수 건수 집계
            daily_strat_counts = {'qty': 0, 'amount': 0, 'percent': 0, 'HTS': 0, 'none': 0}
            strat_stats = {} # { '전략키': {'pnl': 0, 'buy_amt': 0, 'count': 0, 'nm': '명칭'} }
            cond_stats = {} # { '조건식명': {'pnl': 0, 'buy_amt': 0, 'count': 0} }

            for r in processed_data:
                if r.get('buy_qty', 0) > 0:
                    # 전략별 집계
                    s_key = r.get('strat_key', 'none')
                    s_nm = r.get('strat_nm', '--')
                    daily_strat_counts[s_key] = daily_strat_counts.get(s_key, 0) + 1
                    
                    if s_key not in strat_stats:
                        strat_stats[s_key] = {'pnl': 0, 'buy_amt': 0, 'count': 0, 'nm': s_nm}
                    strat_stats[s_key]['pnl'] += r['pnl']
                    strat_stats[s_key]['buy_amt'] += r['buy_amt']
                    strat_stats[s_key]['count'] += 1
                    
                    # [신규] 조건식별 집계 (v3.4)
                    c_name = r.get('cond_name', '직접매매')
                    if c_name not in cond_stats:
                        cond_stats[c_name] = {'pnl': 0, 'buy_amt': 0, 'count': 0}
                    cond_stats[c_name]['pnl'] += r['pnl']
                    cond_stats[c_name]['buy_amt'] += r['buy_amt']
                    cond_stats[c_name]['count'] += 1

            # [신규] 당일 전체 퀀트 지표 계산 (v3.1)
            win_trades = [r for r in processed_data if r['pnl'] > 0]
            loss_trades = [r for r in processed_data if r['pnl'] < 0]
            win_count = len(win_trades)
            loss_count = len(loss_trades)
            win_rate = (win_count / count * 100) if count > 0 else 0
            
            total_profit = sum(r['pnl'] for r in win_trades)
            total_loss = abs(sum(r['pnl'] for r in loss_trades))
            profit_factor = (total_profit / total_loss) if total_loss > 0 else (total_profit if total_profit > 0 else 0)
            
            avg_profit = (total_profit / win_count) if win_count > 0 else 0
            avg_loss = (total_loss / loss_count) if loss_count > 0 else 0
            payoff_ratio = (avg_profit / avg_loss) if avg_loss > 0 else 0
            
            win_prob = win_count / count if count > 0 else 0
            loss_prob = loss_count / count if count > 0 else 0
            expectancy = (win_prob * avg_profit) - (loss_prob * avg_loss)
            
            # MDD (오늘 거래 내역 기준)
            peak = 0
            current_pnl = 0
            mdd = 0
            for r in processed_data:
                current_pnl += r['pnl']
                if current_pnl > peak: peak = current_pnl
                dd = peak - current_pnl
                if dd > mdd: mdd = dd

            # [수정] 샤프 지수 계산 (수익률의 안정성 평가)
            sharpe_ratio = 0
            returns = [r['pnl_rt'] for r in processed_data if r.get('pnl_rt') is not None]
            if len(returns) > 1:
                avg_ret = sum(returns) / len(returns)
                # 편차 제곱의 합 계산
                sum_sq_diff = sum((x - avg_ret) ** 2 for x in returns)
                # 표본 표준편차 (n-1)
                var = sum_sq_diff / (len(returns) - 1)
                std = math.sqrt(var)
                
                if std > 0:
                    # 샤프 지수 = (평균 수익률 - 무위험 수익률) / 표준편차
                    # 단타 매매에서는 무위험 수익률을 0으로 가정
                    sharpe_ratio = avg_ret / std
                else:
                    # 모든 수익률이 동일하여 변동성이 0인 경우 (단, 수익률이 플러스여야 함)
                    sharpe_ratio = 10.0 if avg_ret > 0 else 0.0
            elif len(returns) == 1:
                # 데이터가 하나뿐일 때: 수익이 났다면 최소한의 점수 부여
                sharpe_ratio = 1.0 if returns[0] > 0 else 0.0

            avg_pnl_rt = (total_pnl / abs(total_b_amt) * 100) if abs(total_b_amt) > 100 else 0
            
            # [신규] 최고 수익 시간(Peak Profit Time) 추적
            peak_pnl = 0
            peak_pnl_time = "09:00:00"
            running_pnl = 0
            
            # 시간순으로 정렬된 데이터를 바탕으로 수익 정점 기록
            time_sorted_data = sorted(processed_data, key=lambda x: x['buy_time'])
            for r in time_sorted_data:
                running_pnl += r['pnl']
                if running_pnl > peak_pnl:
                    peak_pnl = running_pnl
                    # [수정] buy_time을 기준으로 하되, 실제로는 매도 시점이 수익 확정 시점이므로 
                    # 리포트에서는 '수익이 발생한 시점'으로 안내
                    peak_pnl_time = r['buy_time'] 

            # 리포트 반환용 stats 딕셔너리 구성
            current_stats = {
                'total_buy': total_b_amt,
                'total_sell': total_s_amt,
                'total_tax': total_tax,
                'total_pnl': total_pnl,
                'count': count,
                'avg_pnl_rt': avg_pnl_rt,
                'daily_strat_counts': daily_strat_counts,
                'win_rate': win_rate,
                'profit_factor': profit_factor,
                'payoff_ratio': payoff_ratio,
                'expectancy': expectancy,
                'mdd': mdd,
                'sharpe_ratio': sharpe_ratio,
                'peak_pnl': peak_pnl,
                'peak_pnl_time': peak_pnl_time,
                'cond_stats': cond_stats, # [신규] 조건식 통계 추가
                'strat_stats': strat_stats # [복구] 매수 전략별 통계 추가
            }

            if sync_only:
                # [v5.5] API 확정 수익으로 실시간 차트 데이터(session_logger) 동기화
                try:
                    from trade_logger import session_logger
                    # [v5.6.5] 리포트용 합계 수익(total_pnl)을 동기화 값으로 사용 (계산 로직 일원화)
                    run_pnl = total_pnl
                    
                    # [v5.6.3] 실시간 타임라인 오염 방지 및 스팸 로그 제거 (정밀도 개선: abs() > 1)
                    if abs(session_logger.cumulative_pnl - run_pnl) > 1:
                        if len(session_logger.pnl_history) <= 1:
                            # 앱 초기 기동 시점에는 과거 흐름을 대략적으로 복원
                            sync_history = [{'time': '09:00:00', 'pnl': 0}]
                            temp_pnl = 0
                            for r in time_sorted_data:
                                temp_pnl += r['pnl']
                                time_str = r['buy_time'].replace('[', '').replace(']', '').replace('전일 ', '').strip()
                                if len(time_str) >= 8 and ':' in time_str:
                                    sync_history.append({'time': time_str[-8:], 'pnl': temp_pnl})
                                else:
                                    sync_history.append({'time': time_str, 'pnl': temp_pnl})
                            session_logger.pnl_history = sync_history
                        else:
                            # 실시간 매매 중 API와 싱크가 어긋났을 때(HTS 강제청산 등) 그 시점부터 새 기준점 추가
                            now_time = datetime.now().strftime("%H:%M:%S")
                            session_logger.pnl_history.append({'time': now_time, 'pnl': run_pnl})
                            
                        session_logger.cumulative_pnl = run_pnl
                        session_logger.save_session()
                        # print(f"🔄 [차트 동기화] API 확정 수익 최신화 완료 (현재: {run_pnl:,}원)") # 사용자 요청으로 스팸 방지 목적 주석 처리
                        
                        # [v5.6.4] 동기화 성공 시 실시간 차트 위젯 갱신 신호 발송
                        if self.on_graph_update:
                            self.on_graph_update()
                except Exception as sync_e:
                    pass
                
                return True

            # [v5.6.5] 일반 리포트 요청 시에도 차트 데이터를 강제로 자동 동기화 (이격 정체 해결)
            try:
                from trade_logger import session_logger
                if abs(session_logger.cumulative_pnl - total_pnl) > 1:
                    now_time = datetime.now().strftime("%H:%M:%S")
                    session_logger.pnl_history.append({'time': now_time, 'pnl': total_pnl})
                    session_logger.cumulative_pnl = total_pnl
                    session_logger.save_session()
                    if self.on_graph_update: self.on_graph_update()
            except: pass

            if summary_only:
                summary_msg = "<b>📝 [ 당일 매매 요약 리포트 ]</b>\n"
                summary_msg += "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                summary_msg += f"🔹 거래종목: {count}건\n"
                # [수정] 수익은 빨간색(#ff4444), 손실은 파란색(#33b5e5)
                pnl_color = "#ff4444" if total_pnl >= 0 else "#33b5e5"
                
                summary_msg += f"🔹 총 매수: {total_b_amt:,}원\n"
                summary_msg += f"🔹 총 매도: {total_s_amt:,}원\n"
                summary_msg += "──────────────────────────────────────────\n"
                summary_msg += f"💸 제세공과: {total_tax:,}원\n"
                summary_msg += f"✨ 실현손익: <font color='{pnl_color}'><b>{total_pnl:+,}원</b></font>\n"
                summary_msg += f"📈 최종수익률: <font color='{pnl_color}'><b>{avg_pnl_rt:+.2f}%</b></font>\n"
                summary_msg += "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                
                if send_telegram:
                    real_tel_send(summary_msg, parse_mode='HTML')
                    print("📢 텔레그램으로 요약 보고서를 전송했습니다.")
                
                tel_send(summary_msg.replace('<b>', '').replace('</b>', ''))
                return True

            display_rows = [] 
            tel_rows = []     
            
            h_line = "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            header = " [시간] [전략] 종목     |  매수액  |  매도액  |  세금  | 손익(수익률) \n"
            
            display_rows.append(h_line + header + h_line)
            tel_rows.append(h_line + header + h_line)

            colors = {'qty': '#ff4444', 'amount': '#00c851', 'percent': '#33b5e5', 'none': '#00ff00', 'HTS': '#8a2be2'}
            for r in processed_data:
                row_color = colors.get(r['strat_key'], '#00ff00')
                bt_str = f"[{r['buy_time']}]"
                if r.get('is_overnight'):
                    bt_str = f"<font color='#ffeb3b'><b>{bt_str}</b></font>" 
                
                st_str = f"[{r['strat_nm']}]"
                
                buy_avg_str = f"{r['buy_avg']:>7,}" if r['buy_avg'] > 0 else f"{'-':>7}"
                buy_qty_str = f"{r['buy_qty']:>3}" if r['buy_qty'] > 0 else f"{'-':>3}"
                buy_amt_str = f"{r['buy_amt']:>8,}" if r['buy_amt'] > 0 else f"{'-':>8}"
                
                row_content = f"{bt_str:<10} {st_str:<6} {r['name']:<10} | {buy_amt_str:>8} | {r['sel_amt']:>8,} | {r['tax']:>5,} | {r['pnl']:>+8,} ({r['pnl_rt']:>+6.2f}%)\n"
                
                row_tel = f"[{r['buy_time']:<8}] {st_str:<6} {r['name']:<10} | {buy_amt_str:>8} | {r['sel_amt']:>8,} | {r['tax']:>5,} | {r['pnl']:>+8,} ({r['pnl_rt']:>+6.2f}%)\n"
                
                display_rows.append(f"<font color='{row_color}'>{row_content}</font>")
                tel_rows.append(row_tel)

            d_ft = "----------------------------------------------------------------------------------\n"
            display_rows.append(d_ft)
            tel_rows.append(d_ft)
            
            summary_str = f"{'TOTAL':<21} {'  ':<6} {'합계':<10} | {total_b_amt:>8,} | {total_s_amt:>8,} | {total_tax:>5,} | {total_pnl:>+8,} ({avg_pnl_rt:>+6.2f}%)\n"
            display_rows.append(summary_str)
            tel_rows.append(summary_str)
            
            display_rows.append(h_line)
            tel_rows.append(h_line)
            
            # [신규] 조건식별 요약 통계 추가 (v3.4)
            if cond_stats:
                s_header = "\n [ 조건식별 누적 통계 ] \n"
                s_h_line = "─────────────────────────────────────────────────────────────────\n"
                display_rows.append(s_header + s_h_line)
                tel_rows.append(s_header + s_h_line)
                
                # 수익금액 기준 내림차순 정렬
                sorted_conds = sorted(cond_stats.items(), key=lambda x: x[1]['pnl'], reverse=True)
                for c_name, c_data in sorted_conds:
                    c_pnl = c_data['pnl']
                    c_rt = (c_pnl / c_data['buy_amt'] * 100) if c_data['buy_amt'] > 0 else 0
                    pnl_color = "#ff4444" if c_pnl >= 0 else "#33b5e5"
                    
                    c_row_html = f" 🔹 {c_name[:12]:<12} : {c_pnl:>+9,}원 ({c_rt:>+6.2f}%) [{c_data['count']}건]\n"
                    c_row_plain = f" 🔹 {c_name[:12]:<12} : {c_pnl:>+9,}원 ({c_rt:>+6.2f}%) [{c_data['count']}건]\n"
                    display_rows.append(f"<font color='{pnl_color}'>{c_row_html}</font>")
                    tel_rows.append(c_row_plain)
                
                display_rows.append(s_h_line)
                tel_rows.append(s_h_line)
                
            # [신규] 매수 전략별 요약 통계 추가 (v4.1.1)
            if strat_stats:
                st_header = "\n [ 매수 전략별 매매현황 ] \n"
                st_h_line = "─────────────────────────────────────────────────────────────────\n"
                display_rows.append(st_header + st_h_line)
                tel_rows.append(st_header + st_h_line)
                
                # 수익금액 기준 내림차순 정렬
                sorted_strats = sorted(strat_stats.items(), key=lambda x: x[1]['pnl'], reverse=True)
                for s_key, s_data in sorted_strats:
                    s_pnl = s_data['pnl']
                    s_rt = (s_pnl / s_data['buy_amt'] * 100) if s_data['buy_amt'] > 0 else 0
                    pnl_color = "#ff4444" if s_pnl >= 0 else "#33b5e5"
                    
                    s_row_html = f" 🔹 {s_data['nm']:<12} : {s_pnl:>+9,}원 ({s_rt:>+6.2f}%) [{s_data['count']}건]\n"
                    s_row_plain = f" 🔹 {s_data['nm']:<12} : {s_pnl:>+9,}원 ({s_rt:>+6.2f}%) [{s_data['count']}건]\n"
                    display_rows.append(f"<font color='{pnl_color}'>{s_row_html}</font>")
                    tel_rows.append(s_row_plain)
                
                display_rows.append(st_h_line)
                tel_rows.append(st_h_line)
                
            # [신규] 상세 리포트에 전체 퀀트 지표 요약 추가
            quant_summary = f"\n📝 [ 오늘 전체 매매 요약 리포트 ]\n"
            quant_summary += f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            quant_summary += f"👑 최고 수익 구간 : {peak_pnl_time} ({peak_pnl:+,}원)\n"
            quant_summary += f"📊 승률 : {win_rate:.1f}% ({win_count}승 {loss_count}패)\n"
            quant_summary += f"💰 PF (수익 지수) : {profit_factor:.2f}\n"
            quant_summary += f"⚖️ 기댓값 : {int(expectancy):,}원 (손익비 {payoff_ratio:.2f})\n"
            quant_summary += f"📉 MDD (최대낙폭) : {int(mdd):,}원\n"
            quant_summary += f"📈 샤프 지수 : {sharpe_ratio:.2f}\n"
            quant_summary += f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            
            display_rows.append(quant_summary)
            tel_rows.append(quant_summary)
            
            if not return_text:
                tel_send("".join(display_rows))
            
            if send_telegram:
                # 텔레그램 전송 시 HTML 파싱 오류 방지를 위해 태그가 섞여있지 않은지 확인
                real_tel_send("".join(tel_rows))
                print("📢 텔레그램으로 상세 보고서 및 요약 통계를 전송했습니다.")
            
                # [v5.6] 기존 Matplotlib 삭제 및 메인 위젯 캡처본(png) 활용 로직
                try:
                    from openpyxl import Workbook
                    from openpyxl.drawing.image import Image as OpenpyxlImage
                    import io
                    
                    df_data = [{
                        '매수시간': r['buy_time'], '매수전략': r['strat_nm'], '조건식': r['cond_name'], 
                        '종목명': r['name'], '종목코드': r['code'], '매수평균가': r['buy_avg'], 
                        '매수수량': r['buy_qty'], '매수금액': r['buy_amt'], '매도평균가': r['sel_avg'], 
                        '매도수량': r['sel_qty'], '매도금액': r['sel_amt'], '세금': r['tax'], 
                        '손익금액': r['pnl'], '수익률(%)': r['pnl_rt']
                    } for r in processed_data]
                    
                    df_data.append({
                        '매수시간': '합계', '매수전략': '-', '조건식': '-', 
                        '종목명': '-', '종목코드': '-', '매수평균가': 0, 
                        '매수수량': 0, '매수금액': total_b_amt, '매도평균가': 0, 
                        '매도수량': 0, '매도금액': total_s_amt, '세금': total_tax, 
                        '손익금액': total_pnl, '수익률(%)': avg_pnl_rt
                    })
                    
                    df = pd.DataFrame(df_data)
                    date_str = datetime.now().strftime("%Y%m%d")
                    
                    import string
                    suffix_list = list(string.ascii_lowercase) 
                    
                    final_filename = f"trade_log_{date_str}.csv"
                    csv_path = os.path.join(self.data_dir, final_filename)
                    
                    if os.path.exists(csv_path):
                        for char in suffix_list:
                            temp_name = f"trade_log_{date_str}_{char}.csv"
                            if not os.path.exists(os.path.join(self.data_dir, temp_name)):
                                final_filename = temp_name
                                csv_path = os.path.join(self.data_dir, final_filename)
                                break
                    
                    df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                    tel_send(f"<font color='#28a745'>📂 당일 매매 일지가 로컬에 저장되었습니다: {final_filename}</font>")
                    
                    # [V5.6] 차트 캡처본 확보 및 HTML 표시
                    png_filename = final_filename.replace('.csv', '.png')
                    png_path = os.path.join(self.data_dir, png_filename)
                    img_buf = io.BytesIO() # 엑셀 첨부 호환을 위한 더미
                    
                    capture_success = False
                    if hasattr(self, 'main_window') and self.main_window:
                        capture_success = self.main_window.capture_profit_graph(png_path)
                    
                    if capture_success and os.path.exists(png_path):
                        # 캡처본을 바이트 스트림(img_buf)에도 로드하여 엑셀에서 쓸 수 있게 함
                        with open(png_path, 'rb') as f:
                            img_buf.write(f.read())
                        img_buf.seek(0)
                        
                        img_url = f"file:///{png_path.replace(chr(92), '/')}" 
                        img_html = f"<br><img src='{img_url}' width='480'/><br>"
                        
                        # [v5.6.5] GUI 전용 고속 HTML 채널Prefix 사용 (필터 우회 및 즉시 렌더링)
                        print(f"DEBUG_HTML_LOG: {img_html}") 
                        
                        display_rows.append(img_html)
                        current_stats['img_html'] = img_html
                        
                        try:
                            real_tel_send_photo(png_path)
                            print("📢 텔레그램으로 차트 사진을 전송했습니다.")
                        except Exception as e:
                            print(f"⚠️ 텔레그램 차트 사진 전송 실패: {e}")
                        
                        # 엑셀 파일 저장 경로 설정 (.xlsx)
                        excel_filename = final_filename.replace('.csv', '.xlsx')
                        excel_path = os.path.join(self.data_dir, excel_filename)
                        
                        # 엑셀에 데이터프레임 기록
                        # xlsxwriter 엔진 대신 openpyxl을 사용하면 이미지를 쉽게 삽입할 수 있음
                        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                            df.to_excel(writer, index=False, sheet_name='매매일지')
                            workbook = writer.book
                            worksheet = writer.sheets['매매일지']
                            
                            # 차트 이미지 삽입 (우측 P2 셀 쯤에 위치)
                            if capture_success and img_buf:
                                try:
                                    img = OpenpyxlImage(img_buf)
                                    worksheet.add_image(img, 'P2')
                                except Exception as img_err:
                                    print(f"⚠️ 이미지를 엑셀에 삽입하지 못했습니다: {img_err}")

                            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                            
                            # --- [신규] 표 내용(매수전략별) 글씨 색상 및 전체적인 디자인 적용 ---
                            thin_border = Border(left=Side(style='thin'), 
                                                 right=Side(style='thin'), 
                                                 top=Side(style='thin'), 
                                                 bottom=Side(style='thin'))

                            header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
                            header_font = Font(color="FFFFFF", bold=True)
                            sum_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                            sum_font = Font(bold=True)
                            
                            font_red = Font(color="E74C3C")      # 적색 (1주)
                            font_green = Font(color="28A745")    # 초록색 (금액)
                            font_blue = Font(color="007BFF")     # 파란색 (비율)
                            font_purple = Font(color="9B59B6")   # 보라색 (HTS)
                            
                            max_row = worksheet.max_row
                            max_col = worksheet.max_column
                            
                            # 1. 헤더 (1행) 디자인
                            for c in range(1, max_col + 1):
                                cell = worksheet.cell(row=1, column=c)
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                                
                            # 2. 합계 줄 (마지막 행) 디자인
                            for c in range(1, max_col + 1):
                                cell = worksheet.cell(row=max_row, column=c)
                                cell.fill = sum_fill
                                cell.font = sum_font
                            
                            # 3. 데이터 줄 (매수전략 조건별로 글씨 색상 적용) 및 전체 테두리
                            for r in range(1, max_row + 1):
                                # r이 데이터 행일 때 글자색 판단을 위해 전략(column=2) 미리 읽기
                                strat_val = ""
                                if 1 < r < max_row:
                                    strat_val = str(worksheet.cell(row=r, column=2).value or "")
                                    
                                t_font = None
                                if '한주' in strat_val: t_font = font_red
                                elif '금액' in strat_val: t_font = font_green
                                elif '비율' in strat_val or '비중' in strat_val: t_font = font_blue
                                elif '수동' in strat_val or 'HTS' in strat_val or '직접' in strat_val: t_font = font_purple
                                
                                for c in range(1, max_col + 1):
                                    cell = worksheet.cell(row=r, column=c)
                                    cell.border = thin_border # 모든 셀 테두리 적용
                                    
                                    # 헤더나 합계 줄이 아닐 때만 폰트 색상 덮어쓰기
                                    if 1 < r < max_row and t_font:
                                        cell.font = t_font
                            
                            # 4. 각 열 너비 데이터에 맞춰 자동 조절
                            for col in worksheet.columns:
                                max_length = 0
                                column_letter = col[0].column_letter
                                for cell in col:
                                    try:
                                        val = str(cell.value) if cell.value is not None else ""
                                        # 한글/영문 차이를 보정하기 위해 euc-kr 인코딩 길이로 추정
                                        length = len(val.encode('euc-kr'))
                                        if length > max_length:
                                            max_length = length
                                    except:
                                        pass
                                # 텍스트 길이에 맞춰 약간의 여백 추가
                                adjusted_width = max_length * 1.1 + 2
                                if adjusted_width > 40: 
                                    adjusted_width = 40
                                worksheet.column_dimensions[column_letter].width = adjusted_width

                            # --- [신규] 리포트 요약 텍스트 추가 (차트 하단 'P35' 부터) ---
                            
                            # 색상 및 스타일 정의
                            bg_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                            white_font = Font(color="FFFFFF", bold=True)
                            red_font = Font(color="FF4444", bold=True)
                            blue_font = Font(color="33B5E5", bold=True)
                            gold_font = Font(color="F1C40F", bold=True)
                            green_font = Font(color="00C851", bold=False)

                            # 차트가 없으면 P2부터 바로 텍스트 출력, 있으면 P38부터 (이미지가 꽤 크기 때문에 겹침 방지)
                            start_row = 38 if capture_success else 2
                            col_p = 16  # 'P' column is 16th

                            def write_cell(r, text, font=white_font, fill=bg_fill):
                                cell = worksheet.cell(row=r, column=col_p)
                                cell.value = text
                                cell.font = font
                                cell.fill = fill
                                cell.alignment = Alignment(vertical='center')
                                return cell
                            
                            r = start_row
                            write_cell(r, f"📝 [ 오늘 전체 매매 종합 리포트 ]")
                            r += 1
                            write_cell(r, f"📅 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", green_font)
                            r += 1
                            
                            total_pnl = current_stats.get('total_pnl', 0)
                            avg_pnl_rt = current_stats.get('avg_pnl_rt', 0)
                            pnl_font = red_font if total_pnl >= 0 else blue_font
                            write_cell(r, f"💰 당일 총손익 : {total_pnl:+,}원 ({avg_pnl_rt:+.2f}%)", pnl_font)
                            r += 1
                            
                            peak_time_str = current_stats.get('peak_pnl_time', '약속된 시간 없음')
                            peak_pnl_val = current_stats.get('peak_pnl', 0)
                            write_cell(r, f"👑 최고 수익 시간 : {peak_time_str} ({peak_pnl_val:+,}원)", gold_font)
                            r += 1
                            
                            wr = current_stats.get('win_rate', 0)
                            pf = current_stats.get('profit_factor', 0)
                            pr = current_stats.get('payoff_ratio', 0)
                            ex = current_stats.get('expectancy', 0)
                            mdd = current_stats.get('mdd', 0)
                            sr = current_stats.get('sharpe_ratio', 0)
                            
                            wr_font = red_font if wr >= 70 else (blue_font if wr <= 40 else white_font)
                            pf_font = red_font if pf >= 2.0 else (blue_font if pf < 1.0 else white_font)
                            
                            write_cell(r, f"📊 승률 : {wr:.1f}%", wr_font)
                            r += 1
                            write_cell(r, f"💰 PF(Profit Factor) : {pf:.2f}", pf_font)
                            r += 1
                            write_cell(r, f"⚖️ 손익비(Payoff Ratio) : {pr:.2f}")
                            r += 1
                            write_cell(r, f"🎯 매매 기댓값 : {int(ex):,}원", gold_font)
                            r += 1
                            write_cell(r, f"📉 MDD(최대낙폭) : {int(mdd):,}원", gold_font)
                            r += 1
                            write_cell(r, f"📈 샤프 지수 : {sr:.2f}")
                            r += 2
                            
                            # [신규] 매수 전략별 매매현황 (v4.1.2)
                            strat_stats = current_stats.get('strat_stats', {})
                            if strat_stats:
                                write_cell(r, "📂 [ 매수 전략별 매매현황 ]")
                                r += 1
                                # 수익금액 기준 내림차순 정렬
                                sorted_strats = sorted(strat_stats.items(), key=lambda x: x[1]['pnl'], reverse=True)
                                for s_key, s_data in sorted_strats:
                                    s_pnl = s_data['pnl']
                                    s_rt = (s_pnl / s_data['buy_amt'] * 100) if s_data['buy_amt'] > 0 else 0
                                    s_font = red_font if s_pnl >= 0 else blue_font
                                    write_cell(r, f" 🔹 {s_data['nm']:<10}: {s_pnl:+,}원 ({s_rt:+.2f}%)", s_font)
                                    r += 1
                                r += 1

                            # [신규] 조건식별 매매현황 (v4.1.2)
                            cond_stats = current_stats.get('cond_stats', {})
                            if cond_stats:
                                write_cell(r, "📂 [ 조건식별 매매현황 ]")
                                r += 1
                                # 수익금액 기준 내림차순 정렬
                                sorted_conds = sorted(cond_stats.items(), key=lambda x: x[1]['pnl'], reverse=True)
                                for c_name, c_data in sorted_conds:
                                    c_pnl = c_data['pnl']
                                    c_rt = (c_pnl / c_data['buy_amt'] * 100) if c_data['buy_amt'] > 0 else 0
                                    c_font = red_font if c_pnl >= 0 else blue_font
                                    write_cell(r, f" 🔹 {c_name[:12]:<12}: {c_pnl:+,}원 ({c_rt:+.2f}%)", c_font)
                                    r += 1
                                r += 1
                            
                            write_cell(r, "📂 [ 오늘 전체 누적 매매현황 ]")
                            r += 1
                            write_cell(r, f"🔹 총매수 : {current_stats.get('total_buy', 0):,}", green_font)
                            r += 1
                            write_cell(r, f"🔹 총매도 : {current_stats.get('total_sell', 0):,}", green_font)
                            r += 1
                            write_cell(r, f"🔹 세금외 : {current_stats.get('total_tax', 0):,}", green_font)
                            r += 1
                            write_cell(r, f"✨ 손익 : {total_pnl:+,}원 ({avg_pnl_rt:+.2f}%)", pnl_font)
                            r += 2
                            
                            # 보유 종목 정보
                            write_cell(r, "📈 [ 현재 보유 종목 ]")
                            r += 1
                            try:
                                account_cache = ACCOUNT_CACHE.get('holdings', {})
                                if account_cache:
                                    for code, s in account_cache.items():
                                        try:
                                            pl_rt = float(s.get('pl_rt', 0))
                                            emoji = "📈" if pl_rt > 0 else "📉"
                                            color_font = red_font if pl_rt > 0 else blue_font
                                            write_cell(r, f"{emoji} {s.get('name', 'N/A')}: {pl_rt:+.2f}% ({int(s.get('pnl', 0)):,}원)", color_font)
                                            r += 1
                                        except:
                                            pass
                                else:
                                    write_cell(r, "현재 보유 중인 종목이 없습니다.", green_font)
                            except Exception as e:
                                write_cell(r, "보유 종목 조회 불가", white_font)
                            
                            # 열 너비 조정 (보고서 텍스트 짤림 방지용)
                            worksheet.column_dimensions['P'].width = 60

                        tel_send(f"<font color='#00d1b2'>📊 <b>그래프가 포함된 엑셀 보고서 생성 완료:</b> {excel_filename}</font>")
                        print(f"✅ 엑셀 및 차트 파일 생성 완료: {excel_path}")
                except Exception as ex_err:
                    print(f"❌ 엑셀 및 그래프 생성 오류: {ex_err}")
                
                except Exception as save_err: 
                    print(f"❌ csv 저장 오류: {save_err}")

            # [신규] 결과 반환 로직 확장 (report 및 sync에서 사용)
            report_text = "".join(display_rows)
            if return_stats and return_text:
                return report_text, current_stats
            if return_stats:
                return current_stats
            if return_text:
                return report_text
            
            # [추가] sync 전용 데이터 리스트 반환
            return processed_data

        except Exception as e:
            print(f"❌ today 오류: {e}")
            tel_send(f"❌ today 오류: {e}")
            return None if return_text else False

    async def tpr(self, number):
        if self.update_setting('take_profit_rate', float(number)):
            tel_send(f"✅ 익절 기준: {number}%")

    async def slr(self, number):
        rate = -abs(float(number))
        if self.update_setting('stop_loss_rate', rate):
            tel_send(f"✅ 손절 기준: {rate}%")

    async def brt(self, number):
        if self.update_setting('buy_ratio', float(number)):
            tel_send(f"✅ 매수 비중: {number}%")

    async def condition(self, number=None, quiet=False):
        try:
            await self.stop(set_auto_start_false=False, quiet=quiet)
            if number is not None:
                if self.update_setting('search_seq', str(number)):
                    tel_send(f"✅ 조건식 {number}번으로 변경")
                    if MarketHour.is_market_open_time(): await self.start()
                    return True
            token = self.token if self.token else self.get_token()
            cond_list = await asyncio.wait_for(get_condition_list(token), timeout=5.0)
            if cond_list:
                cond_list.sort(key=lambda x: int(x[0]))
                for c in cond_list:
                    self.rt_search.condition_map[str(c[0])] = c[1]

                if not quiet:
                    msg = "📋 [조건식 목록]\n"
                    for c in cond_list: msg += f"• {c[0]}: {c[1]}\n"
                    log_and_tel(msg)
            return True
        except: 
            if not quiet: log_and_tel("❌ 목록 조회 실패")

    def update_setting(self, key, value):
        return self.update_settings_batch({key: value})

    def update_settings_batch(self, updates_dict):
        """여러 설정을 한 번에 안전하게 업데이트 (레이스 컨디션 방지)"""
        try:
            settings = {}
            if os.path.exists(self.settings_file):
                with open(self.settings_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
            
            settings.update(updates_dict)
                
            with open(self.settings_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, indent=4, ensure_ascii=False)
            return True
        except Exception as e:
            print(f"❌ 설정 저장 실패: {e}")
            return False

    async def help(self):
        help_msg = """🤖 [명령어 가이드]
• start / stop : 시작 및 중지
• r (또는 report) : 잔고 확인
• auto {번호} : {번호}번 부터 시퀀스 가동 (0은 중지)
• condition {번호} : 조건식 변경
• tpr / slr / brt : 익절/손절/비중 설정
• today 옵션 : 당일 매매 일지 조회
  - today : 시간순
  - today jun : 전략순 (매수전략)
  - today sic : 조건식순 (검색식명)
  - today son : 손익순 (손익금액)
• AI NEWS ALL : 시장 전반 AI 브리핑
• AI NEWS {키워드} : 종목/테마 AI 뉴스 분석
• exp : 로그 데이터 폴더(LogData) 열기
• voice on/off : 매수 시 음성(TTS) 켜기/끄기
• beep on/off : 모든 비프음 소리 켜기/끄기
• tel on/off : 텔레그램 전체 메세지 켜기/끄기
• tel log on/off : 매수/매도 알림만 켜기/끄기 (리포트는 유지)
• tel today : 텔레그램으로 매매 요약 리포트 전송
• clr : 로그 화면 초기화 (GUI 전용)
• log : 현재 로그를 .txt 파일로 저장 (GUI 전용)
• msg {메세지} : 텔레그램 메세지 직접 전송"""
        tel_send(help_msg)

    async def process_command(self, text):
        cmd_full = text.strip()
        cmd = cmd_full.lower()
        
        if cmd == 'start':
            if self.on_start_request: self.on_start_request()
            else: await self.start()
        elif cmd == 'stop':
            if self.on_stop_request: self.on_stop_request()
            else: await self.stop(True)
        elif cmd in ['report', 'r']: await self.report()
        elif cmd.startswith('auto'):
            parts = cmd_full.split()
            idx = 1
            if len(parts) > 1:
                try: idx = int(parts[1])
                except: idx = 1
            
            if self.on_auto_sequence:
                self.on_auto_sequence(idx)
            else:
                tel_send("ℹ️ auto 명령어는 GUI 환경에서만 작동합니다.")
        elif cmd == 'condition': await self.condition()
        elif cmd.startswith('condition '): await self.condition(cmd_full.split()[1])
        elif cmd.startswith('tpr '): await self.tpr(cmd_full.split()[1])
        elif cmd.startswith('slr '): await self.slr(cmd_full.split()[1])
        elif cmd.startswith('brt '): await self.brt(cmd_full.split()[1])
        elif cmd == 'clr':
            if self.on_clear_logs: self.on_clear_logs()
            else: tel_send("ℹ️ clr 명령어는 GUI 환경에서만 작동합니다.")
        elif cmd == 'voice on':
            self.update_setting('voice_guidance', True)
            log_and_tel("🔊 음성 안내가 활성화되었습니다.")
        elif cmd == 'voice off':
            self.update_setting('voice_guidance', False)
            log_and_tel("🔇 음성 안내가 비활성화되었습니다.")
        elif cmd == 'beep off':
            self.update_setting('beep_sound', False)
            log_and_tel("🔕 비프음이 비활성화되었습니다.")
        elif cmd == 'tel on':
            self.update_setting('tel_on', True)
            log_and_tel("🟢 텔레그램 전체 메시지 전송이 활성화되었습니다.")
        elif cmd == 'tel off':
            self.update_setting('tel_on', False)
            log_and_tel("🔴 텔레그램 전체 메시지 전송이 비활성화되었습니다.")
        elif cmd == 'tel log on':
            self.update_setting('tel_log_on', True)
            log_and_tel("🔔 텔레그램 매수/매도 알림이 활성화되었습니다.")
        elif cmd == 'tel log off':
            self.update_setting('tel_log_on', False)
            log_and_tel("🔕 텔레그램 매수/매도 알림이 비활성화되었습니다. (리포트는 전송됨)")
        elif cmd == 'log':
            if self.on_request_log_file: self.on_request_log_file()
            else: tel_send("ℹ️ log 명령어는 GUI 환경에서만 작동합니다.")
        elif cmd == 'print' or cmd == 'msg':
            tel_send(f"❓ {cmd} 뒤에 메세지를 입력해주세요. (예: {cmd} 안녕하세요)")
        elif cmd.startswith('print '):
            await asyncio.get_event_loop().run_in_executor(None, log_and_tel, cmd_full[6:].strip())
        elif cmd.startswith('msg '):
            await asyncio.get_event_loop().run_in_executor(None, log_and_tel, cmd_full[4:].strip())
        elif cmd.startswith('tel_send '):
            await asyncio.get_event_loop().run_in_executor(None, log_and_tel, cmd_full[9:].strip())
        elif cmd == 'refresh_conditions': 
            await self.rt_search.refresh_conditions(self.token)
        elif cmd == 'exp':
            if os.path.exists(self.data_dir):
                os.startfile(self.data_dir)
                log_and_tel("📂 로그 데이터 폴더(LogData)를 엽니다.")
            else:
                log_and_tel("❌ 로그 폴더가 존재하지 않습니다.")
        elif cmd == 'open_config':
            if self.on_open_config: self.on_open_config()
            else: tel_send("ℹ️ 설정창 열기는 GUI 환경에서만 작동합니다.")
        elif cmd == 'open_ai_settings':
            if self.on_open_ai_settings: self.on_open_ai_settings()
            else: tel_send("ℹ️ AI 설정창 열기는 GUI 환경에서만 작동합니다.")
        elif cmd.startswith('ai 뉴스') or cmd.startswith('ai news'):
            parts = cmd_full.split()
            if len(parts) >= 2:
                # [신규 v5.3.0] AI NEWS ALL 또는 AI NEWS [키워드] 처리
                # parts[0] = AI, parts[1] = NEWS/뉴스, parts[2:] = 키워드
                keyword = " ".join(parts[2:]).strip() if len(parts) > 2 else ""
                
                is_all = (keyword.lower() == 'all')
                
                if is_all:
                    tel_send("⏳ <b>전체 시장 뉴스 수집 및 AI 브리핑 시작...</b> (잠시만 기다려주세요)", parse_mode='HTML')
                elif keyword:
                    tel_send(f"⏳ '{keyword}' 최신 뉴스 검색 및 AI 분석 시작... (최대 10초 소요)")
                else:
                    tel_send("❓ 키워드를 입력해주세요. (예: AI NEWS 삼성전자 / AI NEWS ALL)")
                    return

                from news_sniper import run_news_sniper
                loop = asyncio.get_event_loop()
                result_msg = await loop.run_in_executor(None, run_news_sniper, keyword, is_all)
                
                if result_msg:
                    # HTML 태그가 포함되어 있으므로 parse_mode='HTML'을 사용해 전송
                    await loop.run_in_executor(None, real_tel_send, result_msg, 'HTML')
                    
                    # [신규] GUI가 켜져 있다면 팝업창으로도 결과 전송
                    if self.on_news_result:
                        self.on_news_result(result_msg)
                    
                    # [V5.7.1] AI 시장 브리핑(is_all) 결과는 음성으로도 즉시 브리핑하도록 연동
                    if is_all and self.on_ai_voice_response:
                        # HTML 태그 제거 및 TTS용 클린 텍스트 생성
                        clean_briefing = re.sub(r'<[^>]*>', '', result_msg)
                        # 인사말 추가하여 더욱 친근하게 브리핑
                        voice_msg = f"자기야! 오늘 시장 브리핑 들려줄게. {clean_briefing}"
                        self.on_ai_voice_response(voice_msg)
            else:
                tel_send("❓ 키워드를 함께 입력해주세요. (예: AI NEWS ALL / AI NEWS 삼성전자)")
        elif cmd == 'help': await self.help()
        elif cmd.startswith('tel today'):
            sub_raw = cmd_full[4:].strip() 
            is_rev = sub_raw.endswith('-')
            
            parts = sub_raw.lower().split()
            sub_cmd = 'default'
            if len(parts) > 1:
                sub_part = parts[1].replace('-', '')
                if sub_part: sub_cmd = sub_part
            
            is_summary = (sub_raw.lower() == 'today')
            
            if sub_cmd == 'sic': await self.today(sort_mode='sic', is_reverse=is_rev, send_telegram=True)
            elif sub_cmd == 'jun': await self.today(sort_mode='jun', is_reverse=is_rev, send_telegram=True)
            elif sub_cmd == 'son': await self.today(sort_mode='son', is_reverse=is_rev, send_telegram=True)
            else: await self.today(summary_only=is_summary, is_reverse=is_rev, send_telegram=True)

        elif cmd.startswith('today'):
            parts = cmd.split()
            sub_cmd = 'default'
            is_rev = False
            
            full_text = cmd
            is_rev = full_text.endswith('-')
            
            if len(parts) > 1:
                sub_part = parts[1].replace('-', '')
                if sub_part: sub_cmd = sub_part
            elif ' ' not in full_text and len(full_text) > 5:
                sub_part = full_text[5:].replace('-', '')
                if sub_part: sub_cmd = sub_part
                
            if sub_cmd in ['default', 'today']: await self.today(is_reverse=is_rev, send_telegram=False)
            elif sub_cmd == 'sic': await self.today(sort_mode='sic', is_reverse=is_rev, send_telegram=False)
            elif sub_cmd == 'jun': await self.today(sort_mode='jun', is_reverse=is_rev, send_telegram=False)
            elif sub_cmd == 'son': await self.today(sort_mode='son', is_reverse=is_rev, send_telegram=False)
            else: await self.today(is_reverse=is_rev, send_telegram=False) 
        elif cmd == 'close_bet':
            # [신규 v5.7.28] 종가 베팅 추천 (3시 이후 권장)
            now = datetime.now()
            if now.hour < 15:
                 tel_send("⚠️ 아직 종가 베팅을 분석하기엔 이른 시간이야. (보통 15:00 이후 권장)")
            
            tel_send("⏳ <b>오늘 시장을 복기하고 종가 베팅 종목을 분석 중이야...</b> (잠시만 기다려줘)", parse_mode='HTML')
            
            from news_sniper import run_news_sniper
            loop = asyncio.get_event_loop()
            result_msg = await loop.run_in_executor(None, run_news_sniper, None, False, True)
            
            if result_msg:
                await loop.run_in_executor(None, real_tel_send, result_msg, 'HTML')
                if self.on_news_result:
                    self.on_news_result(result_msg)
                
                if self.on_ai_voice_response:
                    clean_msg = re.sub(r'<[^>]*>', '', result_msg)
                    voice_msg = f"자기야! 오늘 종가 베팅 추천 종목 분석 결과가 나왔어. {clean_msg}"
                    self.on_ai_voice_response(voice_msg)

                # [신규 v6.1.16] AI 추천 종목 자동 1주 매수 (종가 베팅 전용)
                from news_sniper import extract_target_stocks_from_msg, get_stock_code_by_name
                from buy_stock import fn_kt10000 as buy_stock
                
                target_names = extract_target_stocks_from_msg(result_msg)
                if target_names:
                    log_and_tel(f"🚀 <b>AI 추천 종목 자동 매수 기동:</b> {', '.join(target_names)} (각 1주씩)")
                    token = await loop.run_in_executor(None, self.get_valid_token)
                    
                    for name in target_names:
                        code = await loop.run_in_executor(None, get_stock_code_by_name, name)
                        if code:
                            # 시장가(3)로 1주 매수 시도
                            ret_code, ret_msg = await loop.run_in_executor(None, buy_stock, code, 1, 0, '3', 'N', '', token)
                            if ret_code == '0':
                                log_and_tel(f"✅ <b>[자동매수 성공]</b> {name}({code}) 1주 시장가 주문 완료!")
                            else:
                                log_and_tel(f"❌ <b>[자동매수 실패]</b> {name}: {ret_msg}")
                        else:
                            log_and_tel(f"⚠️ {name} 종목의 코드를 찾을 수 없어 매수를 생략했어.")
                else:
                    log_and_tel("ℹ️ 자동 매수할 명시적인 종목 태그([BUY:...])를 찾지 못했어.")

                # [신규 v6.1.17] 종가 베팅 4번 항목: KipoStock 관점 분석
                from trade_logger import session_logger
                perspective_list = await loop.run_in_executor(None, session_logger.get_kipostock_perspective, token)
                if perspective_list:
                    # GUI 시그널 발생 (메인 스레드에서 다이얼로그 팝업)
                    self.signals.perspective_signal.emit(perspective_list)
                    
                    # 텔레그램으로도 간략히 보고
                    names = [f"{item['name']}({item['peak_rt']:.1f}%)" for item in perspective_list]
                    log_and_tel(f"🎯 <b>[KipoStock 관점 추천]:</b> {', '.join(names)} 종목이 포착됐어! (GUI에서 체크!")

        elif cmd.startswith('buy_one_share '):
            # [신규 v6.1.17] 특정 종목 1주 시장가 즉시 매수
            # args[0], args[1] 대신 cmd_full에서 분리
            parts = cmd_full.split()
            if len(parts) >= 3:
                code, name = parts[1], parts[2]
                from buy_stock import fn_kt10000 as buy_stock
                loop = asyncio.get_event_loop()
                token = await loop.run_in_executor(None, self.get_valid_token)
                ret_code, ret_msg = await loop.run_in_executor(None, buy_stock, code, 1, 0, '3', 'N', '', token)
                if ret_code == '0':
                    log_and_tel(f"✅ <b>[KipoStock 관점 주문성공]</b> {name}({code}) 1주 시장가 체결 완료!")
                else:
                    log_and_tel(f"❌ <b>[KipoStock 관점 주문실패]</b> {name}: {ret_msg}")
            else:
                log_and_tel("⚠️ buy_one_share 명령어 형식이 잘못되었습니다. (예: buy_one_share 005930 삼성전자)")

        elif cmd == 'ai_news_brief_simple':
            # [신규 v5.7.32] 시장 간편 브리핑 (종목 중심)
            tel_send("⏳ <b>바쁜 자기를 위해 핵심 종목만 쏙쏙 뽑아서 분석 중이야...</b> (잠시만 기다려줘)", parse_mode='HTML')
            
            from news_sniper import run_news_sniper
            loop = asyncio.get_event_loop()
            result_msg = await loop.run_in_executor(None, run_news_sniper, None, False, False, True)
            
            if result_msg:
                await loop.run_in_executor(None, real_tel_send, result_msg, 'HTML')
                if self.on_news_result:
                    self.on_news_result(result_msg)
                
                if self.on_ai_voice_response:
                    clean_msg = re.sub(r'<[^>]*>', '', result_msg)
                    voice_msg = f"자기야! 시장 핵심 요약 들려줄게. {clean_msg}"
                    self.on_ai_voice_response(voice_msg)
        elif cmd == 'export_today_excel':
            # [신규 v6.1.12 Hotfix] 당일 매매 내역을 엑셀/CSV로 수동 내보내기
            await self.today(send_telegram=True)
            log_and_tel("📊 <b>당일 매매 내역 엑셀/CSV 생성이 완료되었습니다.</b>", parse_mode='HTML')
        elif cmd == 'sync_google_drive':
            # [신규 v6.1.12 Hotfix] 당일 매매 내역을 구글 시트에 동기화
            await self.sync_to_google_drive()
        elif cmd.startswith('query_sql'):
            # [신규 v6.1.25] 로컬 DB(PostgreSQL) 데이터 분석 액션
            question = cmd_full[9:].strip()
            log_and_tel(f"🔍 <b>[AI DB 분석 요청]:</b> '{question}'... (로컬 DB 분석을 시작할게!)")
            # 실제 구현은 DB 연결 설정 후 진행 (현재는 인지 단계)
            tel_send("💡 (안내) 현재 PostgreSQL 서버 연동 준비 중이야. 조금만 더 학습하면 완벽한 데이터 통계를 내줄게!")
        elif cmd.startswith('market_signals'):
            # [신규 v6.1.25] Alpha Vantage 금융 지표 조회 액션
            target = cmd_full[14:].strip()
            log_and_tel(f"📡 <b>[AI 지표 조회 요청]:</b> '{target}'... (Alpha Vantage 데이터를 가져오는 중!)")
            # 실제 구현은 Alpha Vantage API 연동 모듈 호출
            tel_send(f"📊 {target}에 대한 기술적 지표(RSI/MACD)는 현재 분석 준비 중이야. API 키 연동 후 곧 보여줄게!")
        elif cmd.startswith('analyze_sheets'):
            # [신규 v6.1.25] 구글 시트 데이터 분석 액션
            target = cmd_full[14:].strip()
            log_and_tel(f"📊 <b>[AI 시트 분석 요청]:</b> '{target}'... (구글 시트 수익 데이터를 읽는 중!)")
            # 실제 구현은 gspread로 시트 데이터 로드 후 제미나이 재분석
            tel_send("📑 구글 시트의 과거 매매 패턴을 입체적으로 분석해서 곧 리포트로 알려줄게, 자기야!")
 
        else:
            # [NEW V5.4.0] 알 수 없는 명령어는 자연어로 간주하고 제미나이 AI 비서에게 넘김
            try:
                from gemini_bot import process_natural_language
                loop = asyncio.get_event_loop()
                # [v6.3.0+] AI 비서 작동 여부를 시각적으로 보여주기 위해 로그 유지
                log_and_tel(f"💬 <b>[AI 비서 요청]:</b> '{text}'... (제미나이가 고민 중이야!)", parse_mode='HTML')
                
                ai_response = await loop.run_in_executor(None, process_natural_language, text)
                
                if ai_response["type"] == "action":
                    log_and_tel(f"⚡ <b>[AI 명령 실행]:</b> {ai_response['cmd']}")
                    await self.process_command(ai_response["cmd"])
                elif ai_response["type"] == "msg":
                    tel_send(f"💬 [AI 비서] {ai_response['msg']}")
                    # GUI의 speak_text 호출을 위한 훅
                    if self.on_ai_voice_response:
                        self.on_ai_voice_response(ai_response["msg"])
                else:
                    # 에러 메시지 출력
                    tel_send(f"⚠️ {ai_response['msg']}")
            except Exception as e:
                tel_send(f"❓ 명령어 또는 자연어 해석 오류: {text} ({e})")

    async def sync_to_google_drive(self):
        """오늘 매매 데이터를 구글 시트에 동기화합니다."""
        try:
            log_and_tel("⏳ <b>구글 드라이브 동기화 시작...</b> (잠시만 기다려줘)", parse_mode='HTML')
            
            # 1. 오늘 매매 데이터 가져오기 (processed_data 리스트 반환)
            data_list = await self.today(sync_only=True)
            if not data_list or not isinstance(data_list, list):
                log_and_tel("📭 동기화할 매매 내역이 없네, 자기야!")
                return

            # 2. 구글 시트 동기화 모듈 호출
            try:
                from google_drive_sync import GoogleSheetSync
                sync_tool = GoogleSheetSync()
                success, msg = await sync_tool.sync_all_trades_today(self)
                
                if success:
                    log_and_tel(f"✅ <b>구글 드라이브 동기화 완료!</b><br>{msg}", parse_mode='HTML')
                else:
                    log_and_tel(f"❌ <b>동기화 실패:</b> {msg}", parse_mode='HTML')
            except ImportError as e:
                log_and_tel(f"❌ <b>동기화 모듈 오류:</b> google_drive_sync.py가 없거나 gspread 라이브러리가 부족해. ({e})")
            except Exception as e:
                log_and_tel(f"❌ <b>동기화 오류:</b> {e}")
                
        except Exception as e:
            log_and_tel(f"❌ <b>구글 드라이브 동기화 프로세스 오류:</b> {e}")
